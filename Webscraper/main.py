import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
from datetime import datetime

def main(nex):
    
    now = datetime.now()
    dt_string = str(now.strftime("%d/%m/%Y %H:%M:%S"))
    
    #Imports data from Google Finance
    req1 = requests.get("https://www.google.com/finance/quote/UBSG:SWX")
    soup = BeautifulSoup(req1.content, "html.parser" )

    res1 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res1.get_text()
    output1 = c_output.replace("CHF", "")
    
    req2 = requests.get("https://www.google.com/finance/quote/SCHN:SWX")
    soup = BeautifulSoup(req2.content, "html.parser" )

    res2 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res2.get_text()
    output2 = c_output.replace("CHF", "")
    
    req3 = requests.get("https://www.google.com/finance/quote/MONC:BIT")
    soup = BeautifulSoup(req3.content, "html.parser" )

    res3 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res3.get_text()
    output3 = c_output.replace("€", "")
    
    req4 = requests.get("https://www.google.com/finance/quote/KNEBV:HEL")
    soup = BeautifulSoup(req4.content, "html.parser" )

    res4 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res4.get_text()
    output4 = c_output.replace("€", "")
    
    req5 = requests.get("https://www.google.com/finance/quote/GCO:BME")
    soup = BeautifulSoup(req5.content, "html.parser" )

    res5 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res5.get_text()
    output5 = c_output.replace("€", "")
    
    req6 = requests.get("https://www.google.com/finance/quote/TCSG:MCX")
    soup = BeautifulSoup(req6.content, "html.parser" )

    res6 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res6.get_text()
    output6 = c_output.replace("₽", "")
    
    req7 = requests.get("https://www.google.com/finance/quote/ADM:LON")
    soup = BeautifulSoup(req7.content, "html.parser" )

    res7 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res7.get_text()
    output7 = c_output.replace("GBX", "")
    
    req8 = requests.get("https://www.google.com/finance/quote/IMB:LON")
    soup = BeautifulSoup(req8.content, "html.parser" )

    res8 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res8.get_text()
    output8 = c_output.replace("GBX", "")
    
    req9 = requests.get("https://www.google.com/finance/quote/NXT:LON")
    soup = BeautifulSoup(req9.content, "html.parser" )

    res9 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res9.get_text()
    output9 = c_output.replace("GBX", "")
    
    req10 = requests.get("https://www.google.com/finance/quote/GOOGL:NASDAQ")
    soup = BeautifulSoup(req10.content, "html.parser" )

    res10 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res10.get_text()
    output10 = c_output.replace("$", "")
    
    req11 = requests.get("https://www.google.com/finance/quote/AMZN:NASDAQ")
    soup = BeautifulSoup(req11.content, "html.parser" )

    res11 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res11.get_text()
    output11 = c_output.replace("$", "")
    
    req12 = requests.get("https://www.google.com/finance/quote/BRK.B:NYSE")
    soup = BeautifulSoup(req12.content, "html.parser" )

    res12 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res12.get_text()
    output12 = c_output.replace("$", "")
    
    req13 = requests.get("https://www.google.com/finance/quote/CACC:NASDAQ")
    soup = BeautifulSoup(req13.content, "html.parser" )

    res13 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res13.get_text()
    output13 = c_output.replace("$", "")
    
    req14 = requests.get("https://www.google.com/finance/quote/FB:NASDAQ")
    soup = BeautifulSoup(req14.content, "html.parser" )

    res14 = soup.find(jsname="ip75Cb", class_="kf1m0")

    c_output = res14.get_text()
    output14= c_output.replace("$", "")
    

    #Exports the data to the Excel spreadsheet
    wb = load_workbook('quotes.xlsx')
    ws = wb.active
    
    ws['A'+nex].value = dt_string
    ws['B'+nex].value = output1
    ws['C'+nex].value = output2
    ws['D'+nex].value = output3
    ws['E'+nex].value = output4
    ws['F'+nex].value = output5
    ws['G'+nex].value = output6
    ws['H'+nex].value = output7
    ws['I'+nex].value = output8
    ws['J'+nex].value = output9
    ws['K'+nex].value = output10
    ws['L'+nex].value = output11
    ws['M'+nex].value = output12
    ws['N'+nex].value = output13
    ws['O'+nex].value = output14
   
    wb.save('quotes.xlsx')
    
    #Writes to the text file
    textfile = open("A-num.txt", "r")
    txtplus1 = int(textfile.read())+1
    textfile.close
    textfile= open("A-num.txt", "w")
    textfile.write(str(txtplus1))
    textfile.close

#Programm loop  
while True :
    rn = str(datetime.now().time())
    if rn >= "10:00:00.000000" and rn <= "10:00:02.000000":
    
        time.sleep(2)
        textfile = open("A-num.txt", "r")
        lex = textfile.read()
        textfile.close
        main(str(lex))
